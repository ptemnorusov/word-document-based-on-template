import os
import sys
import datetime as dt
import glob
import time

import pyodbc
import xlwings as xw
import openpyxl

from docxtpl import DocxTemplate, InlineImage


context = {
    'dogovor_number': '',
    'dogovor_date': '',
    'doveritel_companyname': 'ООО РОМАШКА',
    'doveritel_position': '',
    'doveritel_fullname': '',
    'doveritel_authority': '',
    'doveritel_full_string1': '',
    'doveritel_full_string2': '',
    'doveritel_full_string3': '',
    'akt_number': '',
    'faktura_number': '',
    'akt_date': '',
    'akt_filename': '',
    'case1_id': '',  ############# 1
    'invoice1_id': '',
    'inv1_date': '',
    'inv1_amount': 0.00,
    'inv1_amount_money': '',
    'inv1_vat': 0.00,
    'inv1_vat_money': '',
    'inv1_paid_amount': 0.00,
    'inv1_paid_amount_money': '',
    'inv1_amount_topay': 0.00,
    'inv1_service0_text': '',
    'inv1_service0_amount': 0,
    'inv1_service0_vat': 0.00,
    'inv1_service1_text': '',
    'inv1_service1_amount': 0,
    'inv1_service1_vat': 0.00,
    'inv1_service2_text': '',
    'inv1_service2_amount': 0,
    'inv1_service2_vat': 0.00,
    'inv1_service3_text': '',
    'inv1_service3_amount': 0,
    'inv1_service3_vat': 0.00,
    'case2_id': '',  ############## 2
    'invoice2_id': '',
    'inv2_date': '',
    'inv2_amount': 0.00,
    'inv2_amount_money': '',
    'inv2_vat': 0.00,
    'inv2_vat_money': '',
    'inv2_paid_amount': 0.00,
    'inv2_paid_amount_money': '',
    'inv2_amount_topay': 0.00,
    'inv2_service0_text': '',
    'inv2_service0_amount': 0,
    'inv2_service0_vat': 0.00,
    'inv2_service1_text': '',
    'inv2_service1_amount': 0,
    'inv2_service1_vat': 0.00,
    'inv2_service2_text': '',
    'inv2_service2_amount': 0,
    'inv2_service2_vat': 0.00,
    'inv2_service3_text': '',
    'inv2_service3_amount': 0,
    'inv2_service3_vat': 0.00,
    'case3_id': '',  ############# 3
    'invoice3_id': '',
    'inv3_date': '',
    'inv3_amount': 0.00,
    'inv3_amount_money': '',
    'inv3_vat': 0.00,
    'inv3_vat_money': '',
    'inv3_paid_amount': 0.00,
    'inv3_paid_amount_money': '',
    'inv3_amount_topay': 0.00,
    'inv3_service0_text': '',
    'inv3_service0_amount': 0,
    'inv3_service0_vat': 0.00,
    'inv3_service1_text': '',
    'inv3_service1_amount': 0,
    'inv3_service1_vat': 0.00,
    'inv3_service2_text': '',
    'inv3_service2_amount': 0,
    'inv3_service2_vat': 0.00,
    'inv3_service3_text': '',
    'inv3_service3_amount': 0,
    'inv3_service3_vat': 0.00,
    'inv_all_amount': 0.00,
    'inv_all_amount_money': '',
    'inv_all_paid_amount': 0.00,
    'inv_all_paid_amount_money': '',
    'inv_all_vat': 0.00,
    'inv_all_vat_money': '',
    'inv_all_amount_topay': 0.00,
    'inv_all_amount_topay_money': '',
    'image': '',
    'edo': ''}


def gather_excel_data(invoice_id, i):
    """
    Функция для сбора необходимых для акта данных из эксель файла
    :param invoice_id:  #Номер счета для поиска
    :param i:           #Порядковый номер для услуг по счету

    """
    # Читаем файл бухгалтерскую таблицу, ищем нужную нам строку со счетом
    for row in ws.iter_rows(min_row=1, min_col=9, max_col=9):
        for cell in row:
            if cell.value == invoice_id:
                our_row = row[0].row
                break
    # Добавляем в список значения из строки в эксель файле
    for row in ws.iter_rows(min_row=our_row, max_row=our_row):
        for cell in row:
            ourline.append(cell.value)

    # Наполняем наш словарь соответствующими данными, с соответствующими условиями,
    # конвертируем данные в нужные для отображения форматы
    context[f'invoice{i}_id'] = invoice_id
    context['dogovor_number'] = str(ourline[1])
    context['akt_date'] = str(dt.datetime.strftime(ourline[18], u'%d-%m-%Y'))
    print(f"Так, кажется, я нашел нужный счет {invoice_id}, давай посмотрим, какие там услуги...")
    if ourline[12]:
        if 'руб' not in str(ourline[22]):
            context[f'inv{i}_amount'] = round(float(str(ourline[25])), 2)
            context[f'inv{i}_amount_money'] = "{:,.2f}".format(context[f'inv{i}_amount']).replace(',', ' ')
            context[f'inv{i}_vat'] = 0.00
            context[f'inv{i}_vat_money'] = "{:,.2f}".format(context[f'inv{i}_vat']).replace(',', ' ')
            print('Ага, счет в валюте, НДС не считаем, правильно?')
        else:
            context[f'inv{i}_amount'] = round(float(str(ourline[21])), 2)
            context[f'inv{i}_amount_money'] = "{:,.2f}".format(context[f'inv{i}_amount']).replace(',', ' ')
            context[f'inv{i}_vat'] = 0.00
            context[f'inv{i}_vat_money'] = "{:,.2f}".format(context[f'inv{i}_vat']).replace(',', ' ')
            print('Ага, счет в рублях, а, значит, пошлина - НДС не считаем')
    else:
        context[f'inv{i}_amount'] = round(float(str(ourline[21])), 2)
        context[f'inv{i}_amount_money'] = "{:,.2f}".format(context[f'inv{i}_amount']).replace(',', ' ')
        context[f'inv{i}_vat'] = round((float(str(ourline[21])) / 1.20 * 0.20), 2)
        context[f'inv{i}_vat_money'] = "{:,.2f}".format(context[f'inv{i}_vat']).replace(',', ' ')
        print('Ух ты, наши спецы постарались - наши услуги, считаем НДС:)')
        if ourline[17]:
            if 'ав' in ourline[17]:
                if ourline[17].split(' ')[-1].isdigit():
                    context['faktura_number'] = ourline[17].split(' ')[-1]
                elif ourline[17].split(' ')[-2].isdigit():
                    context['faktura_number'] = ourline[17].split(' ')[-2]
                else:
                    sheet.range("E2").value = "Счет-фактура не найдена"
            else:
                context['faktura_number'] = ourline[17].strip()

    if not ourline[24]:
        context[f'inv{i}_paid_amount'] = format(0.00, '.2f')
        context[f'inv{i}_paid_amount_money'] = "{:,.2f}".format(0.00).replace(',', ' ')
        print(
            "Ой, Надежда, нам не заплатили... Надеюсь, акт выполненных работ напомнит им, что мы не бесплатно работаем :)")
    else:
        context[f'inv{i}_paid_amount'] = round(float(str(ourline[25])), 2)
        context[f'inv{i}_paid_amount_money'] = "{:,.2f}".format(context[f'inv{i}_paid_amount']).replace(',', ' ')
        print(f"Вижу оплату {context[f'inv{i}_paid_amount_money']} руб., все в порядке")
    if ourline[0]:
        if 'ЭДО' in ourline[0]:
            context['edo'] = "ЭДО"
        else:
            context['edo'] = ""
    if ourline[9]:
        context[f'inv{i}_date'] = str(dt.datetime.strftime((ourline[9]), u'%d-%m-%Y'))
    else:
        context[f'inv{i}_date'] = ""

    print("Еще немного магии и все будет готово. Только терпение.. :)")

    context[f'inv{i}_amount_topay'] = float(context[f'inv{i}_amount']) - float(context[f'inv{i}_paid_amount'])
    context[f'inv{i}_amount_topay_money'] = "{:,.2f}".format(context[f'inv{i}_amount_topay']).replace(',', ' ')

    context['inv_all_amount'] = float(context['inv_all_amount']) + float(context[f'inv{i}_amount'])
    context['inv_all_amount_money'] = "{:,.2f}".format(context['inv_all_amount']).replace(',', ' ')

    context['inv_all_vat'] = float(context['inv_all_vat']) + float(context[f'inv{i}_vat'])
    context['inv_all_vat_money'] = "{:,.2f}".format(context['inv_all_vat']).replace(',', ' ')

    context['inv_all_paid_amount'] = float(context['inv_all_paid_amount']) + float(context[f'inv{i}_paid_amount'])
    context['inv_all_paid_amount_money'] = "{:,.2f}".format(context['inv_all_paid_amount']).replace(',', ' ')

    context['inv_all_amount_topay'] = float(context['inv_all_amount_topay']) + float(context[f'inv{i}_amount_topay'])
    context['inv_all_amount_topay_money'] = "{:,.2f}".format(context['inv_all_amount_topay']).replace(',', ' ')

    # Собираем строку для шапки акта
    if ourline[1]:
        context['dogovor_number'] = ourline[1]

    if ourline[34]:
        context['doveritel_position'] = ourline[34]
    else:
        context['doveritel_position'] = "__________________"

    if ourline[35]:
        context['doveritel_fullname'] = ourline[35]
    else:
        context['doveritel_fullname'] = "_______________"

    context['doveritel_authority'] = ourline[36]

    if context['doveritel_authority']:
        if 'Устав' in context['doveritel_authority']:
            context[
                'doveritel_full_string'] = f"{context['doveritel_position']} {context['doveritel_companyname']} {context['doveritel_fullname']}, действующий на основании Устава"
        else:
            context[
                'doveritel_full_string'] = f"{context['doveritel_position']} {context['doveritel_companyname']} {context['doveritel_fullname']}, действующий на основании доверенности {context['doveritel_authority']}"
    else:
        context[
            'doveritel_full_string'] = f"__________________________________________________ {context['doveritel_companyname']} ______________________________________________"

    print("C Экселем закончили.. Теперь берем данные из Патриции")


def gather_sql_data(invoice_id, i):
    """
    Собираем данные по актам из БД MSSQL
    c соответствующими параметрами подключения и номера счета
    :param invoice_id:          # Номер счета для поиска в базе
    :param i:                   # Индекс для наполнения словаря
    :return:
    """

    # Реквизиты подключения
    sqlserver = 'tcp:DBSERVEROIPADDRESS'
    sqlusername = 'DBUSER'
    sqldatabase = 'DBNAME'
    sqlpassword = 'DBUSERPASSWORD'
    cnxn = pyodbc.connect(
        'DRIVER={SQL Server Native Client 11.0};SERVER=' + sqlserver + ';DATABASE=' + sqldatabase + ';UID=' + sqlusername + ';PWD=' + sqlpassword)
    cursor = cnxn.cursor()

    # Looking for INVOICE DATE, ACTOR ID by INVOICE ID
    sqlquery1 = f"""
    SELECT INVOCIE_DATE,ACTOR_ID, CASE_ID FROM dbo.TABLE1 WHERE INVOICE_ID = {invoice_id} """
    cursor.execute(sqlquery1)
    sqldata1 = cursor.fetchall()

    # Выбираем нужные данные из результата запроса в БД
    actor_id = sqldata1[0][1]

    # Делаем запрос к другой таблице в БД. С джоинами не удалось разобраться
    # Получаем полное наименование компании
    sqlquery2 = f"""
    SELECT TOP (1) NAME_PREFIX, NAME1, NAME2, NAME3 FROM dbo.TABLE2 where name_id = {actor_id} and CURRENT_ONE = 1 ORDER BY FIELD
    """
    cursor.execute(sqlquery2)
    sqldata2 = cursor.fetchall()

    # Собираем строку с названием компании
    xstr = lambda s: s or ""
    context['doveritel_companyname'] = xstr(sqldata2[0][0]) + " " + xstr(sqldata2[0][1]) + " " + xstr(sqldata2[0][2]) + " " + xstr(sqldata2[0][3])

    # Чистим строку от запрещенных символов для имени файла
    characters_to_remove = '"()»«'
    for characters in characters_to_remove:
        context['akt_filename'] = context['doveritel_companyname'].replace(characters, "").replace('"', "").replace(
            '  ', '').replace('«', "").replace("»", "").replace("^", "")

    # Делаем третий запрос. Достаем данные по счетам. Текст услуги, количество, стоимость, НДС
    sqlquery3 = f"""
    SELECT	INVOICE_LINE_TEXT, INVOICE_AMOUNT, INVOICE_VAT_AMOUNT 		
    FROM dbo.TABLE3 where INVOICE_ID = {invoice_id}
    """

    cursor.execute(sqlquery3)
    sqldata3 = cursor.fetchall()

    # Наполняем словарь данными из запроса
    s = 0
    for [invoice_line_text, invoice_amount, invoice_vat_amount] in sqldata3:
        context[f'inv{i}_service' + str(s) + '_text'] = invoice_line_text
        s = s + 1
    cnxn.close()
    print("Хм.. В Патриции данных немного... А имя то какое.. Что ж, давай запишем все в наш акт.")


# Ищем и открываем файл со счетом.
def open_invoice_file(invoice_id):
    filename = glob.glob(fr"PATH_WHERE_INVOICES_ARE_LOCATED\*{invoice_id}*.docx", recursive=True)[0]
    os.startfile(filename.strip("'"))


def main():
    # Читаем книгу из которой запустили скрипт (модуль xlwings)
    # Visual Basic Модуль Make_an_akt запускается кнопкой на листе книги
    xw.Book("AKT.xlsm").set_mock_caller()
    wb = xw.Book.caller()
    sheet = wb.sheets[0]

    #Читаем какие данные ввел пользователь для работы
    wbdata1 = sheet.range('B3').options(dict, expand='table', numbers=int).value

    # Чистим диалоговое пространство с пользователем.
    for row_to_clear in range(2, 11):
        sheet.range(f"E{row_to_clear}").value = ""
        sheet.range(f"I{row_to_clear}").value = ""

    # Открыть бухгалтерскую таблицу (только для чтения)
    buhfile = (r'PATH_WHERE_ACCOUNTANT_EXCELFILE_IS_LOCATED.xlsx')
    ws = openpyxl.load_workbook(buhfile, read_only=True, data_only=True, keep_links=False).active
    print("Привет, Надежда, давай сделаем пару актов)")


    # Для каждого счета, введенного пользователем, запустим соответствующие функции и наполним словарь данными.
    for a in range(1, 4):       # До 3-х счетов за раз
        ourline = []
        invoice_id = wbdata1[f'inv{a}_number']

        # Проверяем, ввели ли данные
        if invoice_id:
            gather_sql_data(invoice_id, a)
            gather_excel_data(invoice_id, a)
            open_invoice_file(invoice_id)
            balovstvo()

    # Если в таблице был номер договора, найти все акты с этим номером,
    # посчитать, сколько актов было сделано в этом году и узнать номер следующего акта
    if context['dogovor_number']:
        akt_number = glob.glob(fr"PATH_WHERE_OTHER_AKTS_ARE_LOCATED\*{context['dogovor_number']}*.docx")
        try:
            context['akt_number'] = int(len(akt_number)) + 1
            sheet.range(f"E3").value = f"Номер акта: {context['akt_number']}"
        except:
            sheet.range(f"E3").value = "Вычислить номер следующего акта не удалось"

    # Сформировать имя файла акта и сохранить изменения в файле
    output_filename = rf"PATH_WHERE_TO_SAVE_AKT_TEMPORARILY/{context['akt_filename']}-({context['dogovor_number']})-{context['edo']}-2022-{context['akt_number']}.docx"

    # Если контрагент не подключен к ЭДО, загрузить соответствующий футер акта
    if not context['edo']:
        print(context['edo'])
        context['image'] = akt.new_subdoc('FOOTER.docx')
    
    os.chdir(sys.path[0])
    # Открыть файл шаблона документа
    akt = DocxTemplate('akt-template.docx')

    # Проливаем словарь с данными в файл шаблона  
    akt.render(context)
    akt.save(output_filename)

    # Ждем секунду, чтобы акт открылся последним. Все ведь ради него затевалось
    time.sleep(1)
    os.startfile(output_filename)

    # Пробуем открыть файл со счетом-фактурой, если услуги подразумевают НДС
    faktura_filename = glob.glob(
        rf"PATH_WHERE_SCHET_FAKTURY_ARE_LOCATED\*{context['faktura_number']}*.xls")
    # Специфика такова, что может быть 1 или 2 счет фактуры. 
    # Если что то идет не так, прекращаем процесс открытия с-ф.
    if len(faktura_filename) < 3:
        for faktura in faktura_filename:
            os.startfile(faktura)
    else:
        sheet.range("E2").value = "Я нашел слишком много счетов-фактур.. Открывать не буду"

    # Подтверждаем окончание выполнения скрипта.
    sheet.range("E9").value = "Готово!"

if __name__ == "__main__":
    main()